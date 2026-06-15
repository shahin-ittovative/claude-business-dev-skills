# scripts/export_xlsx.py
"""Clone the PRE-BRD reference workbook and fill Answer cells from a payload.

Formulas, sample columns, styling, and the Executive Summary control panel are
preserved by cloning the workbook file on disk and writing only into cells the
cell map marks as writable.
"""
from __future__ import annotations

import json
import shutil
from copy import copy
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
REFERENCE = ROOT / "reference" / "PRE-BRD-v1.1.xlsx"
CELL_MAP = ROOT / "reference" / "cell-map.json"


class ExportError(Exception):
    pass


def load_cell_map(path: Path = CELL_MAP) -> dict:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def clone_workbook(dst: Path, src: Path = REFERENCE):
    shutil.copyfile(src, dst)
    return openpyxl.load_workbook(dst)


def write_answers(ws, answers: dict, allowed: set) -> None:
    for coord, value in answers.items():
        if coord not in allowed:
            raise ExportError(f"{ws.title}!{coord} is not a writable Answer cell")
        ws[coord] = value


def expand_variable_rows(ws, spec: dict, needed: int) -> dict:
    """Ensure the sheet has `needed` data rows. Insert extra rows after the last
    data row, cloning the template row's style and formula-column patterns, then
    rewrite aggregate formulas to span the new data range.

    Returns {"first": int, "last": int} of the data range after expansion.
    """
    first = spec["first_row"]
    template = spec["template_row"]
    available = spec["last_row"] - first + 1
    extra = max(0, needed - available)

    if extra:
        insert_at = spec["last_row"] + 1
        ws.insert_rows(insert_at, amount=extra)
        for offset in range(extra):
            r = insert_at + offset
            for col in range(1, ws.max_column + 1):
                s = ws.cell(row=template, column=col)
                d = ws.cell(row=r, column=col)
                if s.has_style:
                    d._style = copy(s._style)
            for col, pattern in spec.get("formula_columns", {}).items():
                ws[f"{col}{r}"] = pattern.format(r=r)

    last = first + needed - 1
    for agg in spec.get("aggregates", []):
        cell = f"{agg['column']}{agg['row'] + extra}"
        ws[cell] = agg["pattern"].format(first=first, last=last)
    return {"first": first, "last": last}


def set_control_panel(wb, switch: dict | None) -> None:
    if not switch:
        return
    wb[switch["sheet"]][switch["cell"]] = switch["value"]


def clear_all_answers(wb, cmap: dict) -> None:
    """Blank every whitelisted Answer cell across ALL mapped sheets so the
    reference workbook's example values can never survive into the output —
    not even in a sheet the payload omits. Formula and sample cells are never
    in the whitelist, so they are untouched.
    """
    for sheet, spec in cmap.items():
        if sheet.startswith("_") or sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for coord in spec.get("answer_cells", []):
            ws[coord] = None
        var = spec.get("variable_rows")
        if var:
            for col in var.get("answer_columns", []):
                for r in range(var["first_row"], var["last_row"] + 1):
                    ws[f"{col}{r}"] = None


def scrub_em_dashes(wb) -> None:
    """Replace the em dash (U+2014) with a hyphen in every string cell of the
    output workbook. The reference template's own guidance and READ-ONLY sample
    text contains em dashes; cloning carries them into the deliverable. House
    style forbids the em dash, so scrub the generated file (the reference
    workbook itself is never modified).
    """
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and "—" in c.value:
                    c.value = c.value.replace("—", "-")


def export(payload: dict, dst: Path, *, src: Path = REFERENCE, cell_map: dict | None = None) -> Path:
    cmap = cell_map if cell_map is not None else load_cell_map()
    wb = clone_workbook(Path(dst), src=src)

    # Wipe every mapped sheet's answer cells first, so any sheet the payload
    # omits comes out clean-blank rather than carrying the reference example.
    clear_all_answers(wb, cmap)

    for sheet, block in payload.get("sheets", {}).items():
        if sheet not in cmap:
            raise ExportError(f"Sheet '{sheet}' missing from cell map")
        if sheet not in wb.sheetnames:
            raise ExportError(f"Sheet '{sheet}' missing from workbook")
        ws = wb[sheet]
        spec = cmap[sheet]
        allowed = set(spec.get("answer_cells", []))

        var = spec.get("variable_rows")
        if var:
            needed = block.get("needed_rows", var["last_row"] - var["first_row"] + 1)
            rng = expand_variable_rows(ws, var, needed)
            for col in var.get("answer_columns", []):
                for r in range(rng["first"], rng["last"] + 1):
                    allowed.add(f"{col}{r}")

        write_answers(ws, block.get("answers", {}), allowed)

    if payload.get("control_panel"):
        set_control_panel(wb, cmap.get("_control_panel"))

    scrub_em_dashes(wb)
    wb.save(Path(dst))
    return Path(dst)
