# scripts/tests/test_export_xlsx.py
from copy import copy

import openpyxl
import pytest

import export_xlsx as ex


def test_clone_preserves_formula_and_value(tmp_path):
    src = tmp_path / "src.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 3
    ws["A2"] = "=A1*2"
    wb.save(src)

    dst = tmp_path / "out.xlsx"
    cloned = ex.clone_workbook(dst, src=src)

    assert dst.exists()
    assert cloned.active["A2"].value == "=A1*2"
    assert cloned.active["A1"].value == 3


def test_write_answers_writes_allowed_and_rejects_others(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ex.write_answers(ws, {"D4": 1000, "D5": 2}, allowed={"D4", "D5", "D6"})
    assert ws["D4"].value == 1000
    assert ws["D5"].value == 2

    with pytest.raises(ex.ExportError):
        ex.write_answers(ws, {"D8": 5}, allowed={"D4", "D5", "D6"})


def test_expand_variable_rows_extends_formulas_and_aggregate(synthetic_efas):
    wb = openpyxl.load_workbook(synthetic_efas)
    ws = wb["EFAS"]
    spec = {
        "template_row": 2,
        "first_row": 2,
        "last_row": 6,
        "answer_columns": ["B", "D", "E"],
        "formula_columns": {"F": "=D{r}*E{r}"},
        "aggregates": [{"column": "F", "row": 7, "pattern": "=SUM(F{first}:F{last})"}],
    }
    rng = ex.expand_variable_rows(ws, spec, needed=7)

    assert rng == {"first": 2, "last": 8}
    assert ws["F7"].value == "=D7*E7"
    assert ws["F8"].value == "=D8*E8"
    assert ws["F9"].value == "=SUM(F2:F8)"


def test_expand_variable_rows_no_insert_when_enough(synthetic_efas):
    wb = openpyxl.load_workbook(synthetic_efas)
    ws = wb["EFAS"]
    spec = {
        "template_row": 2, "first_row": 2, "last_row": 6,
        "answer_columns": ["B"], "formula_columns": {"F": "=D{r}*E{r}"},
        "aggregates": [{"column": "F", "row": 7, "pattern": "=SUM(F{first}:F{last})"}],
    }
    rng = ex.expand_variable_rows(ws, spec, needed=4)
    assert rng == {"first": 2, "last": 5}
    assert ws["F7"].value == "=SUM(F2:F5)"


def test_set_control_panel_sets_switch(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Executive Summary"
    ex.set_control_panel(wb, {"sheet": "Executive Summary", "cell": "C5", "value": "YES"})
    assert wb["Executive Summary"]["C5"].value == "YES"

    ex.set_control_panel(wb, None)
    assert wb["Executive Summary"]["C5"].value == "YES"


def _synthetic_src(tmp_path):
    wb = openpyxl.Workbook()
    rice = wb.active
    rice.title = "RICE Framework"
    rice["D8"] = "=(D4*D5*D6)/D7"  # protected formula must survive
    es = wb.create_sheet("Executive Summary")
    es["C5"] = "NO"
    src = tmp_path / "ref.xlsx"
    wb.save(src)
    return src


def test_export_end_to_end(tmp_path):
    src = _synthetic_src(tmp_path)
    cell_map = {
        "_control_panel": {"sheet": "Executive Summary", "cell": "C5", "value": "YES"},
        "RICE Framework": {"answer_cells": ["D4", "D5", "D6", "D7"]},
    }
    payload = {
        "control_panel": True,
        "sheets": {"RICE Framework": {"answers": {"D4": 1000, "D5": 2, "D6": 0.8, "D7": 5}}},
    }
    dst = tmp_path / "out.xlsx"
    ex.export(payload, dst, src=src, cell_map=cell_map)

    wb = openpyxl.load_workbook(dst)
    assert wb["RICE Framework"]["D4"].value == 1000
    assert wb["RICE Framework"]["D8"].value == "=(D4*D5*D6)/D7"
    assert wb["Executive Summary"]["C5"].value == "YES"


def test_export_rejects_unknown_sheet(tmp_path):
    src = _synthetic_src(tmp_path)
    cell_map = {"RICE Framework": {"answer_cells": ["D4"]}}
    payload = {"sheets": {"Nope": {"answers": {"A1": 1}}}}
    with pytest.raises(ex.ExportError):
        ex.export(payload, tmp_path / "o.xlsx", src=src, cell_map=cell_map)


def test_export_clears_stale_cells_in_sheets_absent_from_payload(tmp_path):
    # A sheet the workbook ships pre-filled but the payload never mentions must
    # come out blank, not carrying the reference example's leftover values.
    wb = openpyxl.Workbook()
    a = wb.active
    a.title = "RICE Framework"
    a["D4"] = 1
    other = wb.create_sheet("Ansoff Matrix")
    other["B2"] = "STALE EXAMPLE"  # whitelisted answer cell, not in payload
    src = tmp_path / "ref.xlsx"
    wb.save(src)

    cell_map = {
        "RICE Framework": {"answer_cells": ["D4"]},
        "Ansoff Matrix": {"answer_cells": ["B2"]},
    }
    payload = {"sheets": {"RICE Framework": {"answers": {"D4": 5}}}}  # Ansoff omitted
    dst = tmp_path / "out.xlsx"
    ex.export(payload, dst, src=src, cell_map=cell_map)

    out = openpyxl.load_workbook(dst)
    assert out["RICE Framework"]["D4"].value == 5
    assert out["Ansoff Matrix"]["B2"].value is None  # stale example cleared


def test_export_clears_stale_answer_cells(tmp_path):
    # The reference workbook ships pre-filled with an example. Any whitelisted
    # answer cell the payload does NOT set must come out blank, not stale.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RICE Framework"
    ws["D4"] = 999  # stale example value
    ws["D5"] = 888  # stale example value
    src = tmp_path / "ref.xlsx"
    wb.save(src)

    cell_map = {"RICE Framework": {"answer_cells": ["D4", "D5", "D6", "D7"]}}
    payload = {"sheets": {"RICE Framework": {"answers": {"D5": 2}}}}
    dst = tmp_path / "out.xlsx"
    ex.export(payload, dst, src=src, cell_map=cell_map)

    out = openpyxl.load_workbook(dst)
    assert out["RICE Framework"]["D5"].value == 2       # written
    assert out["RICE Framework"]["D4"].value is None    # stale value cleared


# NOTE: the `aggregates` path in expand_variable_rows is forward-looking.
# The current reference cell-map.json uses no aggregates (EFAS/IFAS are fixed;
# the variable sheets BCG/OKRs/Roadmap have no total formulas), so this path is
# exercised only by the synthetic fixtures above, not by a real export.
