# scripts/tests/conftest.py
import sys
from pathlib import Path

import openpyxl
import pytest

# Make scripts/ importable as the package root for `import export_xlsx`.
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))


@pytest.fixture
def synthetic_efas(tmp_path):
    """A tiny workbook mimicking a variable-row framework: 5 data rows,
    F=D*E per row, and a SUM total directly below."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EFAS"
    ws["B1"], ws["D1"], ws["E1"], ws["F1"] = "Factor", "Weight", "Rating", "Weighted"
    for r in range(2, 7):
        ws[f"F{r}"] = f"=D{r}*E{r}"
    ws["F7"] = "=SUM(F2:F6)"
    path = tmp_path / "synthetic.xlsx"
    wb.save(path)
    return path
