# scripts/tests/test_cell_map.py
import json
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[2]
REF = ROOT / "reference" / "PRE-BRD-v1.1.xlsx"
CMAP = ROOT / "reference" / "cell-map.json"


@pytest.fixture(scope="module")
def cmap():
    return json.loads(CMAP.read_text(encoding="utf-8"))


@pytest.fixture(scope="module")
def wb():
    return openpyxl.load_workbook(REF)


def _is_formula(ws, coord):
    v = ws[coord].value
    return isinstance(v, str) and v.startswith("=")


def test_every_mapped_sheet_exists(cmap, wb):
    for sheet in cmap:
        if sheet.startswith("_"):
            continue
        assert sheet in wb.sheetnames, f"{sheet} not in workbook"


def test_answer_cells_are_not_formulas(cmap, wb):
    for sheet, spec in cmap.items():
        if sheet.startswith("_"):
            continue
        ws = wb[sheet]
        for coord in spec.get("answer_cells", []):
            assert not _is_formula(ws, coord), f"{sheet}!{coord} is a formula; not writable"


def test_variable_formula_columns_are_formulas_in_template(cmap, wb):
    for sheet, spec in cmap.items():
        if sheet.startswith("_"):
            continue
        var = spec.get("variable_rows")
        if not var:
            continue
        ws = wb[sheet]
        t = var["template_row"]
        for col in var.get("formula_columns", {}):
            assert _is_formula(ws, f"{col}{t}"), f"{sheet}!{col}{t} expected a formula"


def test_control_panel_points_at_real_cell(cmap, wb):
    cp = cmap.get("_control_panel")
    if cp:
        assert cp["sheet"] in wb.sheetnames
        assert "cell" in cp and "value" in cp
