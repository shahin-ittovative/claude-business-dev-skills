# scripts/discover_cells.py
"""Print a per-sheet classification of cells in the reference workbook to help
author cell-map.json. Run: python scripts/discover_cells.py"""
import sys
from pathlib import Path

import openpyxl

REF = Path(__file__).resolve().parent.parent / "reference" / "PRE-BRD-v1.1.xlsx"


def header_for(ws, col):
    for r in range(1, 6):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and v.strip():
            return v.strip().replace("\n", " ")[:40]
    return ""


def main():
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    wb = openpyxl.load_workbook(REF)
    for name in wb.sheetnames:
        ws = wb[name]
        print("=" * 70)
        print(f"SHEET: {name}  ({ws.dimensions})")
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                is_formula = isinstance(c.value, str) and c.value.startswith("=")
                hdr = header_for(ws, c.column)
                kind = "FORMULA" if is_formula else "value"
                print(f"  {c.coordinate:5} [{kind:7}] col='{hdr}' :: {str(c.value)[:50]}")


if __name__ == "__main__":
    main()
